from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl
import xlrd
from openpyxl import Workbook
import pathlib

background = "#0091FF"
framebg = "#EDEDED"
framefg = "#06283D"
cor_topside = "#FF9900"
cor_bgregistro = "#0C4FE8"
fonte = "arial 20 bold"
busca = "arial 16"
busca2 = "arial 13 bold"
texto = "arial 13"


janela = Tk()
janela.title("Registro de Estudantes")
janela.geometry("1250x700+210+100")
janela.config(bg=background)


arquivo = pathlib.Path('Registro_Estudantes.xlsx')
if arquivo.exists():
    pass
else:
    arquivo = Workbook()
    conteudo = arquivo.active
    conteudo['A1'] = "Registro Nº"
    conteudo['B1'] = "Data de Nascimento"
    conteudo['C1'] = "Nome"
    conteudo['D1'] = "Serie"
    conteudo['E1'] = "Sala"
    conteudo['F1'] = "Genero"
    conteudo['G1'] = "Habilidades"
    conteudo['H1'] = "Nota 1"
    conteudo['I1'] = "Nota 2"
    conteudo['J1'] = "Nota 3"
    conteudo['K1'] = "Nota 4"
    arquivo.save("Registro_Estudantes.xlsx")

# Funções


def Exit():
    janela.destroy()


def mostrarimagem():
    global nomearquivo
    global imgs
    nomearquivo = filedialog.askopenfilename(initialdir=os.getcwd(), title='Selecione o arquivo da imagem', filetype=(
        ("JPG File", "*jpg"), ("PNG File", "*.png"), ("All Files", "*.txt")))

    imgs = (Image.open(nomearquivo))
    redimensionar_img = imgs.resize((190, 190))
    photo2 = ImageTk.PhotoImage(redimensionar_img)
    pers.config(image=photo2)
    pers.image = photo2


def limpar():
    nome.set('')
    data_nasc.set('')
    salas.set('')
    series.set('')
    turnos.set('')
    genero.set('')
    nota1.set('')
    nota2.set('')
    nota3.set('')
    nota4.set('')

    numero_registro()
    salvarBotao.config(state='normal')


def salvar():
      r1 = registro.get()
      n1 = nome.get()
      d1 = data_nasc.get()
      g1 = genero.get()
      s2 = series.get()
      t1 = turnos.get()
      s1 = salas.get()
      p1 = nota1.get()
      p2 = nota2.get()
      p3 = nota3.get()
      p4 = nota4.get()

      try:
            g1 = genero
      except:
            messagebox.showerror("error", "Selecione um genero!")

      if n1 == "" or d1 == "" or g1 == "" or s2 == "" or t1 == "" or s1 =="" or p1 == "" or p2 == "" or p3 == "" or p4== "":
            messagebox.showerror("error","Preencha todos os dados")
      else:
            file=openpyxl.load_workbook('Registro_Estudantes.xlsx')
            conteudo=file.active()
            conteudo.cell(column=1,row=conteudo.max_row+1,value=r1)
            conteudo.cell(column=2,row=conteudo.max_row,value=n1)
            conteudo.cell(column=3,row=conteudo.max_row,value=d1)
            conteudo.cell(column=4,row=conteudo.max_row,value=g1)
            conteudo.cell(column=5,row=conteudo.max_row,value=s2)
            conteudo.cell(column=6,row=conteudo.max_row,value=t1)
            conteudo.cell(column=7,row=conteudo.max_row,value=s1)
            conteudo.cell(column=8,row=conteudo.max_row,value=p1)
            conteudo.cell(column=9,row=conteudo.max_row,value=p2)
            conteudo.cell(column=10,row=conteudo.max_row,value=p3)
            conteudo.cell(column=11,row=conteudo.max_row,value=p4)

            file.save(r'Registro_Estudantes.xlsx')

def numero_registro():
    file = openpyxl.load_workbook('Registro_Estudantes.xlsx')
    conteudo = file.active
    linha = conteudo.max_row
    max_row_value = conteudo.cell(row=linha, column=1).value

    try:
        registro.set(max_row_value+1)
    except:
        registro.set("1")


def selecionar():
    valor = genero.get()
    if valor == 1:
        gender = 'Masculino'
        print(gender)
    elif valor == 2:
        gender = 'Feminino'
        print(gender)
    else:
        gender = "Outros"
        print(gender)


# Formatação da janela
Label(janela, text="Email: eduardoschrotke@gmail.com", width=10,
      height=3, bg=cor_topside, anchor='e').pack(side=TOP, fill=X)
Label(janela, text="REGISTRO DE ALUNOS", width=10, height=2,
      bg=cor_bgregistro, fg='#fff', font=fonte).pack(side=TOP, fill=X)
buscar = StringVar()
Entry(janela, textvariable=buscar, width=15,
      bd=2, font=busca).place(x=850, y=70)
iconeimagem = PhotoImage(file="images/search.png")
srch = Button(janela, text="Buscar", compound=LEFT,
              image=iconeimagem, width=135, bg=background, font=busca2)
srch.place(x=1080, y=66)
iconeimagem2 = PhotoImage(file='images/user-account.png')
atualizarbotao = Button(janela, image=iconeimagem2, bg=cor_bgregistro)
atualizarbotao.place(x=110, y=64)

# Registro
Label(janela, text='Registro Nº', font=texto,
      fg=framebg, bg=background).place(x=30, y=150)
Label(janela, text='Data', font=texto, fg=framebg,
      bg=background).place(x=500, y=150)
registro = StringVar()
data = StringVar()
entrada_registro = Entry(janela, textvariable=registro,
                         width=15, font='arial 10').place(x=160, y=150)
today = date.today()
d1 = today.strftime("%d/%m/%Y")
entrada_data = Entry(janela, textvariable=data, width=15,
                     font='arial 10').place(x=550, y=150)
numero_registro()

# Informações estudantes
detalhes = LabelFrame(janela, text='Informações do Estudante', font=20, bd=2,
                      width=900, bg=framebg, fg=framefg, height=250, relief=GROOVE).place(x=30, y=200)
Label(detalhes, text='Nome:', font=texto,
      bg=framebg, fg=framefg).place(x=70, y=240)
Label(detalhes, text='Data de Nascimento:', font=texto,
      bg=framebg, fg=framefg).place(x=450, y=240)
Label(detalhes, text='Genero:', font=texto,
      bg=framebg, fg=framefg).place(x=70, y=320)
Label(detalhes, text='Série:', font=texto,
      bg=framebg, fg=framefg).place(x=450, y=320)
Label(detalhes, text='Turno:', font=texto,
      bg=framebg, fg=framefg).place(x=70, y=400)
Label(detalhes, text='Sala:', font=texto,
      bg=framebg, fg=framefg).place(x=450, y=400)

nome = StringVar()
entrada_nome = Entry(detalhes, textvariable=nome, width=20,
                     font='arial10').place(x=200, y=240)

data_nasc = StringVar()
data_nascimento = Entry(detalhes, textvariable=data_nasc, width=20,
                        font='arial10').place(x=620, y=240)

genero = StringVar()
entrada_genero = Entry(detalhes, textvariable=genero, width=20,
                       font='arial10').place(x=200, y=320)

series = StringVar()
entrada_series = Entry(detalhes, textvariable=series, width=20,
                       font='arial10').place(x=620, y=320)

turnos = StringVar()
entrada_turnos = Entry(detalhes, textvariable=turnos, width=20,
                       font='arial10').place(x=200, y=400)

salas = StringVar()
salas1 = Entry(detalhes, textvariable=salas, width=20,
               font='arial10').place(x=620, y=400)


# Informações Notas
detalhes2 = LabelFrame(janela, text='Informações das Notas', font=20, bd=2,
                       width=900, bg=framebg, fg=framefg, height=220, relief=GROOVE).place(x=30, y=470)
Label(detalhes2, text='Nota 01:', font=texto,
      bg=framebg, fg=framefg).place(x=70, y=540)
Label(detalhes2, text='Nota 02:', font=texto,
      bg=framebg, fg=framefg).place(x=70, y=600)
Label(detalhes2, text='Nota 03:', font=texto,
      bg=framebg, fg=framefg).place(x=470, y=540)
Label(detalhes2, text='Nota 04:', font=texto,
      bg=framebg, fg=framefg).place(x=470, y=600)

nota1 = StringVar()
nota10 = Entry(detalhes, textvariable=nota1, width=20,
               font='arial10').place(x=200, y=540)
nota2 = StringVar()
nota20 = Entry(detalhes, textvariable=nota2, width=20,
               font='arial10').place(x=200, y=600)
nota3 = StringVar()
nota30 = Entry(detalhes, textvariable=nota3, width=20,
               font='arial10').place(x=600, y=540)
nota4 = StringVar()
nota40 = Entry(detalhes, textvariable=nota4, width=20,
               font='arial10').place(x=600, y=600)

# Foto
foto = Frame(janela, bd=3, bg="white", width=180,
             height=180, relief=GROOVE).place(x=1000, y=160)
iconeimagem3 = PhotoImage(file='images/photo.png')
pers = Label(foto, bg='white', image=iconeimagem3).place(x=1000, y=160)

# Botao
Button(janela, text='Carregar Foto', width=19, height=2, font=busca2,
       bg='lightblue', command=mostrarimagem).place(x=1000, y=380)
salvarBotao = Button(janela, text='Salvar', width=19, height=2,
                     font=busca2, bg='lightgreen', command=salvar).place(x=1000, y=460)
Button(janela, text='Resetar', width=19, height=2,
       font=busca2, bg='lightpink', command=limpar).place(x=1000, y=540)
Button(janela, text='Sair', width=19, height=2, font=busca2,
       bg='grey', command=Exit).place(x=1000, y=620)


janela.mainloop()
